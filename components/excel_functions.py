import io
from datetime import timedelta

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from .draw_functions import darken_color, rgb_to_hex, hex_to_rgb, lighten_color
from .worktime_functions import calculate_finish_date, get_equipment_total_work_hours


def export_to_excel(jobs_df, equipment_df, start_date, end_date, calendar_dict):
    """Экспорт диаграммы Ганта в Excel файл"""

    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "График загрузки оборудования"

    # Стили границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    # Стили выравнивания
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Рассчитываем общее количество столбцов (1 столбец = 0.25 часа)
    total_columns = 0
    day_columns = {}  # Словарь для хранения информации о днях

    current_date = start_date
    while current_date <= end_date:
        work_hours = calendar_dict.get(current_date.isoformat(), 8)
        is_weekend = (work_hours == 0)
        # Для выходных используем 2 часа (8 столбцов), для рабочих - обычные рабочие часы
        day_columns_count = work_hours * 4 if not is_weekend else 2 * 4  # 2 часа для выходных

        day_columns[current_date] = {
            'start_col': total_columns + 1,
            'end_col': total_columns + day_columns_count,
            'work_hours': work_hours,
            'is_weekend': is_weekend,
            'hours_count': 2 if is_weekend else work_hours  # Количество часов для отображения
        }

        total_columns += day_columns_count
        current_date += timedelta(days=1)

    # Настройка столбцов
    # Первый столбец для названий оборудования
    ws.column_dimensions['A'].width = 25

    # Столбцы для временной шкалы (ширина 1)
    for col in range(2, total_columns + 2):
        ws.column_dimensions[get_column_letter(col)].width = 1

    # ШАПКА ТАБЛИЦЫ - даты
    header_row = 1
    for date, day_info in day_columns.items():
        start_col = day_info['start_col'] + 1  # +1 потому что первый столбец A
        end_col = day_info['end_col'] + 1

        if start_col <= end_col:
            # Объединяем ячейки для дня
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            ws.merge_cells(f'{start_letter}{header_row}:{end_letter}{header_row}')

            # Заполняем дату
            cell = ws[f'{start_letter}{header_row}']
            cell.value = f"{date.strftime('%d.%m')} ({day_info['work_hours']})"
            cell.alignment = center_align
            cell.font = Font(bold=True)
            cell.border = thick_border  # Толстые границы для ячейки с датой

            # Заливка для выходных
            if day_info['is_weekend']:
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # СТРОКИ ОБОРУДОВАНИЯ
    current_row = 2

    for _, equipment in equipment_df.iterrows():

        # Получаем работы для этого оборудования
        equipment_jobs = jobs_df[jobs_df['equipment_name'] == equipment['name']]

        # Рассчитываем общее время работы оборудования в диапазоне
        total_work_hours = get_equipment_total_work_hours(start_date, end_date, equipment_jobs, calendar_dict)

        # Три строки для одного оборудования
        for i in range(3):
            row_height = 8 if i in [0, 2] else 35  # Тонкие строки сверху и снизу, толстая посередине
            ws.row_dimensions[current_row + i].height = row_height

            # Ячейка с названием оборудования (объединенная для трех строк)
            if i == 0:
                ws.merge_cells(f'A{current_row}:A{current_row + 2}')
                equipment_cell = ws[f'A{current_row}']
                equipment_cell.value = f"{equipment['name']}\n({total_work_hours:.2f}ч)"
                equipment_cell.alignment = center_align
                equipment_cell.font = Font(bold=True)
                equipment_cell.border = thick_border
                rgb_color = hex_to_rgb(equipment['type_color'])
                type_color = lighten_color(rgb_color, 0.75)
                equipment_cell.fill = PatternFill(start_color=rgb_to_hex(type_color).strip("#"),
                                                  end_color=rgb_to_hex(type_color).strip("#"), fill_type='solid')

        # ПОДГОТОВКА ДАННЫХ ДЛЯ ОТОБРАЖЕНИЯ РАБОТ
        # Создаем карту занятости для средней строки
        middle_row_occupancy = {}  # key: (start_col, end_col), value: job_data

        for _, job in equipment_jobs.iterrows():
            hour_offset = float(job['hour_offset']) if not pd.isna(job['hour_offset']) else 0.0
            duration_hours = float(job['duration_hours'])
            start_dt = pd.to_datetime(job['start_date'])

            # Рассчитываем расписание работы
            finish_date_str, daily_schedule = calculate_finish_date(
                start_dt.isoformat(),
                duration_hours,
                hour_offset,
            )

            # ДОБАВЛЯЕМ ВЫХОДНЫЕ ДНИ В РАСПИСАНИЕ (ТО ЖЕ САМОЕ ДЛЯ ОТОБРАЖЕНИЯ)
            extended_schedule = []
            work_start_date = pd.to_datetime(job['start_date']).date()
            work_finish_date = pd.to_datetime(finish_date_str).date()

            current_schedule_date = work_start_date
            schedule_index = 0

            while current_schedule_date <= work_finish_date:
                current_date_str = current_schedule_date.isoformat()

                if schedule_index < len(daily_schedule) and daily_schedule[schedule_index][0] == current_date_str:
                    extended_schedule.append(daily_schedule[schedule_index])
                    schedule_index += 1
                else:
                    work_hours = calendar_dict.get(current_date_str, 8)
                    if work_hours == 0:  # Выходной день
                        extended_schedule.append((current_date_str, 0, 0))

                current_schedule_date += timedelta(days=1)

            daily_schedule_with_weekends = extended_schedule

            # Для каждого дня в расширенном расписании
            for day_str, hours, offset in daily_schedule_with_weekends:
                day_date = pd.to_datetime(day_str).date()

                # Проверяем, попадает ли день в диапазон экспорта
                if day_date in day_columns:
                    day_info = day_columns[day_date]

                    # Для выходных дней (hours=0) создаем запись на весь день
                    if hours == 0 and day_info['is_weekend']:
                        start_col = day_info['start_col'] + 1
                        end_col = day_info['end_col'] + 1

                        middle_row_occupancy[(start_col, end_col)] = {
                            'job': job,
                            'day_info': day_info,
                            'is_weekend': True
                        }

                    # Для рабочих дней с работой
                    elif hours > 0:
                        # Рассчитываем столбцы для работы в этот день
                        start_hour_offset = offset
                        end_hour_offset = offset + hours

                        # Для выходных дней ограничиваем 2 часами
                        if day_info['is_weekend']:
                            max_hours = 2
                        else:
                            max_hours = day_info['work_hours']

                        # Ограничиваем рабочими часами дня
                        end_hour_offset = min(end_hour_offset, max_hours)

                        if start_hour_offset < end_hour_offset:
                            # Переводим часы в столбцы (1 столбец = 0.25 часа)
                            start_col_offset = int(start_hour_offset * 4)
                            end_col_offset = int(end_hour_offset * 4)

                            start_col = day_info['start_col'] + 1 + start_col_offset
                            end_col = day_info['start_col'] + 1 + end_col_offset - 1

                            if start_col <= end_col:
                                middle_row_occupancy[(start_col, end_col)] = {
                                    'job': job,
                                    'day_info': day_info,
                                    'is_weekend': False
                                }

        # ЗАПОЛНЕНИЕ ВРЕМЕННОЙ ШКАЛЫ ДЛЯ ОБОРУДОВАНИЯ
        for date, day_info in day_columns.items():
            start_col = day_info['start_col'] + 1
            end_col = day_info['end_col'] + 1

            # СТРОКА 1 и 3 (верхняя и нижняя) - объединяем по часам
            for row_offset in [0, 2]:  # Строки 1 и 3
                row = current_row + row_offset

                # Объединяем кванты в часы (4 кванта = 1 час)
                hours_in_day = day_info['hours_count']

                for hour in range(hours_in_day):
                    hour_start_col = start_col + (hour * 4)
                    hour_end_col = hour_start_col + 3  # 4 кванта на час

                    if hour_start_col <= hour_end_col:
                        # Объединяем 4 кванта в одну ячейку часа
                        start_letter = get_column_letter(hour_start_col)
                        end_letter = get_column_letter(hour_end_col)
                        ws.merge_cells(f'{start_letter}{row}:{end_letter}{row}')

                        # Обращаемся к ПЕРВОЙ ячейке объединенного диапазона
                        hour_cell = ws[f'{start_letter}{row}']

                        # Применяем границы ко всей объединенной ячейке (только вертикальные)
                        _apply_vertical_borders_only(hour_cell, hour_start_col, hour_end_col, start_col, end_col,
                                                     day_info)

                        # Заливка для выходных
                        if day_info['is_weekend']:
                            hour_cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

            # СТРОКА 2 (средняя) - сначала рисуем сетку часов, потом объединяем работы
            middle_row = current_row + 1

            # НАХОДИМ ВСЕ РАБОТЫ В ЭТОМ ДНЕ ДЛЯ СРЕДНЕГО РЯДА
            day_jobs = []
            for (job_start, job_end), job_data in middle_row_occupancy.items():
                if job_start >= start_col and job_end <= end_col:
                    day_jobs.append((job_start, job_end, job_data))

            # Сортируем работы по начальной колонке
            day_jobs.sort(key=lambda x: x[0])

            # Сначала создаем сетку часов во второй строке (как в 1 и 3 строках, но без объединения)
            hours_in_day = day_info['hours_count']
            for hour in range(hours_in_day):
                hour_start_col = start_col + (hour * 4)
                hour_end_col = hour_start_col + 3  # 4 кванта на час

                # Применяем границы к каждому кванту в часе
                for col in range(hour_start_col, hour_end_col + 1):
                    cell = ws[f'{get_column_letter(col)}{middle_row}']

                    # Границы для квантов (только вертикальные)
                    border_parts = {}

                    # Левая граница - толстая если это первый квант часа и первый час дня
                    if col == hour_start_col and hour == 0:
                        border_parts['left'] = Side(style='thick')
                    elif col == hour_start_col:
                        border_parts['left'] = Side(style='thin')
                    else:
                        border_parts['left'] = Side(style=None)

                    # Правая граница - толстая если это последний квант часа и последний час дня
                    if col == hour_end_col and hour == hours_in_day - 1:
                        border_parts['right'] = Side(style='thick')
                    elif col == hour_end_col:
                        border_parts['right'] = Side(style='thin')
                    else:
                        border_parts['right'] = Side(style=None)

                    # Верхняя и нижняя границы - убираем
                    border_parts['top'] = Side(style=None)
                    border_parts['bottom'] = Side(style=None)

                    cell.border = Border(**border_parts)

            # ОБРАБОТКА ВЫХОДНЫХ И РАБОЧИХ ДНЕЙ
            if day_info['is_weekend']:
                # Для выходных дней - создаем объединенную ячейку на весь день
                start_letter = get_column_letter(start_col)
                end_letter = get_column_letter(end_col)
                merge_range = f'{start_letter}{middle_row}:{end_letter}{middle_row}'
                ws.merge_cells(merge_range)

                weekend_cell = ws[f'{start_letter}{middle_row}']

                # Ищем работу, которая включает этот выходной день
                weekend_job = None
                for job_start, job_end, job_data in day_jobs:
                    if job_data.get('is_weekend', False):
                        weekend_job = job_data
                        break

                if weekend_job:
                    # Есть работа в этот выходной - используем приглушенный цвет заказа

                    dark_color = darken_color(hex_to_rgb(weekend_job['job']['order_color']), 0.3)
                    fill_color_hex = rgb_to_hex(dark_color).strip("#")

                    # job_text = weekend_job['job']['order_name']
                    # if len(job_text) > 15:
                    #     job_text = job_text[:15] + '...'
                    # weekend_cell.value = job_text
                    # weekend_cell.alignment = center_align
                    # weekend_cell.font = Font(size=8, bold=True, color='FFFFFF')
                else:
                    # Нет работ - серая заливка
                    fill_color_hex = 'F0F0F0'

                weekend_cell.fill = PatternFill(start_color=fill_color_hex, end_color=fill_color_hex, fill_type='solid')
                _apply_vertical_borders_only(weekend_cell, start_col, end_col, start_col, end_col, day_info)

            else:
                # Для рабочих дней - создаем отдельные ячейки для работ
                for job_start, job_end, job_data in day_jobs:
                    if not job_data.get('is_weekend', False) and job_start <= job_end:  # Только рабочие дни
                        start_letter = get_column_letter(job_start)
                        end_letter = get_column_letter(job_end)

                        merge_range = f'{start_letter}{middle_row}:{end_letter}{middle_row}'
                        ws.merge_cells(merge_range)

                        job_cell = ws[f'{start_letter}{middle_row}']

                        job_text = job_data['job']['order_name']
                        if len(job_text) > 15:
                            job_text = job_text[:15] + '...'

                        try:
                            job_cell.value = job_text
                            job_cell.alignment = center_align
                            job_cell.font = Font(size=8, bold=True)
                        except Exception as e:
                            pass

                        color_hex = job_data['job']['order_color'].lstrip('#')
                        job_cell.fill = PatternFill(
                            start_color=color_hex,
                            end_color=color_hex,
                            fill_type='solid'
                        )

                        if color_hex in ['FFFFFF', 'FFFF00', '00FFFF']:
                            job_cell.font = Font(size=8, bold=True, color='000000')
                        else:
                            job_cell.font = Font(size=8, bold=True, color='FFFFFF')

                        _apply_vertical_borders_only(job_cell, job_start, job_end, start_col, end_col, day_info)

        current_row += 3

    # ОБВОДИМ ВСЮ ОБЛАСТЬ ДИАГРАММЫ ТОЛСТЫМИ ВНЕШНИМИ ГРАНИЦАМИ
    _apply_outer_thick_borders(ws, total_columns, 1, current_row - 1)

    for i, (_, equipment) in enumerate(equipment_df.iterrows()):
        _apply_outer_thick_borders(ws, total_columns, 2 + i * 3, 5 + i * 3 - 1)

    # Сохраняем файл в буфер
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def _apply_vertical_borders_only(cell, cell_start_col, cell_end_col, day_start_col, day_end_col, day_info):
    """Применяет только вертикальные границы к ячейке (без горизонтальных)"""
    border_parts = {}

    # Левая граница - толстая только если это первая ячейка дня
    if cell_start_col == day_start_col:
        border_parts['left'] = Side(style='thick')
    else:
        border_parts['left'] = Side(style='thin')

    # Правая граница - толстая только если это последняя ячейка дня
    if cell_end_col == day_end_col:
        border_parts['right'] = Side(style='thick')
    else:
        border_parts['right'] = Side(style='thin')

    # Верхняя и нижняя границы - убираем (нет границ)
    border_parts['top'] = Side(style=None)
    border_parts['bottom'] = Side(style=None)

    cell.border = Border(**border_parts)


def _apply_outer_thick_borders(ws, total_columns, first_row, last_row):
    """Применяет толстые внешние границы ко всей области диаграммы"""

    # Верхняя граница (шапка с датами)
    for col in range(1, total_columns + 2):
        col_letter = get_column_letter(col)
        cell = ws[f'{col_letter}{first_row}']
        current_border = cell.border
        new_border = Border(
            left=current_border.left,
            right=current_border.right,
            top=Side(style='thick'),  # Толстая верхняя граница
            bottom=current_border.bottom
        )
        cell.border = new_border

    # Нижняя граница
    for col in range(1, total_columns + 2):
        col_letter = get_column_letter(col)
        cell = ws[f'{col_letter}{last_row}']
        current_border = cell.border
        new_border = Border(
            left=current_border.left,
            right=current_border.right,
            top=current_border.top,
            bottom=Side(style='thick')  # Толстая нижняя граница
        )
        cell.border = new_border

    # Левая граница (столбец A)
    for row in range(1, last_row + 1):
        cell = ws[f'A{row}']
        current_border = cell.border
        new_border = Border(
            left=Side(style='thick'),  # Толстая левая граница
            right=current_border.right,
            top=current_border.top,
            bottom=current_border.bottom
        )
        cell.border = new_border

    # Правая граница (последний столбец)
    last_col_letter = get_column_letter(total_columns + 1)
    for row in range(1, last_row + 1):
        cell = ws[f'{last_col_letter}{row}']
        current_border = cell.border
        new_border = Border(
            left=current_border.left,
            right=Side(style='thick'),  # Толстая правая граница
            top=current_border.top,
            bottom=current_border.bottom
        )
        cell.border = new_border
